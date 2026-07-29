from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

GASTOS_COLUMNS = [
    "Año", "Trimestre", "Tipo", "Fecha factura", "Número factura", "Proveedor", "NIF",
    "Concepto", "Dirección", "Código Postal", "Modelo 303", "Modelo 303 Base imponible",
    "Base imponible", "Tipo IVA", "IVA", "Tipo retención", "Retención modelo 111", "Total",
    "% deducción", "Importe deducible IRPF", "Cuenta", "Tipo de gasto",
]

INGRESOS_COLUMNS = [
    "Año", "Trimestre", "Tipo de factura", "Op intracomunit?", "Número factura", "Fecha factura",
    "Cliente", "CIF", "Dirección", "Base IRPF", "Base imponible EUR", "BI Modelo 303",
    "Tipo IVA", "IVA", "Tipo retención", "Retención", "Suplidos", "Importe IRPF",
    "Importe total sin suplidos", "Importe total", "Epígrafe",
]


def safe_number(value: Any) -> float:
    try:
        if value is None or pd.isna(value) or value == "":
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def year_quarter(date_text: str) -> tuple[str, str]:
    try:
        parsed = datetime.strptime(str(date_text), "%d/%m/%Y")
        return str(parsed.year), f"{((parsed.month - 1) // 3 + 1)}T"
    except Exception:
        return "", ""


def export_date(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.strftime("%d-%m-%y")
        except ValueError:
            continue
    return text


def display_rate(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass
    text = str(value).strip()
    if not text:
        return ""
    return text if "%" in text else f"{text}%"


def to_gastos(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for _, row in df.iterrows():
        year, quarter = year_quarter(row.get("fecha_factura", ""))
        is_social_security = str(row.get("tipo_gasto", "")).strip().lower() == "cuota seguridad social" or str(row.get("concepto", "")).strip().lower() == "cuota seguridad social"
        model_value = str(row.get("modelo_303", "") or "").strip()
        model_base = row.get("base_imponible") if model_value.lower() in {"sí", "si"} else None
        accounting_concept = "Cuota seguridad social" if is_social_security else (row.get("tipo_gasto", "") or row.get("concepto", ""))
        records.append({
            "Año": year,
            "Trimestre": quarter,
            "Tipo": "Débito" if is_social_security else row.get("tipo_documento", ""),
            "Fecha factura": export_date(row.get("fecha_factura", "")),
            "Número factura": "" if is_social_security else row.get("numero_factura", ""),
            "Proveedor": row.get("proveedor", ""),
            "NIF": row.get("nif", ""),
            "Concepto": accounting_concept,
            "Dirección": row.get("direccion", ""),
            "Código Postal": row.get("codigo_postal", ""),
            "Modelo 303": "no" if is_social_security else ("sí" if model_value.lower() in {"sí", "si"} else "no" if model_value.lower() == "no" else ""),
            "Modelo 303 Base imponible": model_base,
            "Base imponible": row.get("base_imponible"),
            "Tipo IVA": display_rate(row.get("tipo_iva", "")),
            "IVA": row.get("iva"),
            "Tipo retención": display_rate(row.get("tipo_retencion", "")),
            "Retención modelo 111": row.get("retencion"),
            "Total": row.get("total"),
            "% deducción": "100.00%" if is_social_security else row.get("porcentaje_deduccion", ""),
            "Importe deducible IRPF": row.get("total") if is_social_security else row.get("importe_deducible_irpf"),
            "Cuenta": row.get("cuenta", ""),
            "Tipo de gasto": row.get("tipo_gasto", ""),
        })
    return pd.DataFrame(records, columns=GASTOS_COLUMNS)


def to_ingresos(df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for _, row in df.iterrows():
        year, quarter = year_quarter(row.get("fecha_factura", ""))
        total = row.get("total")
        records.append({
            "Año": year,
            "Trimestre": quarter,
            "Tipo de factura": row.get("tipo_documento", ""),
            "Op intracomunit?": "",
            "Número factura": row.get("numero_factura", ""),
            "Fecha factura": export_date(row.get("fecha_factura", "")),
            "Cliente": row.get("proveedor", ""),
            "CIF": row.get("nif", ""),
            "Dirección": row.get("direccion", ""),
            "Base IRPF": "",
            "Base imponible EUR": row.get("base_imponible"),
            "BI Modelo 303": row.get("base_imponible"),
            "Tipo IVA": display_rate(row.get("tipo_iva", "")),
            "IVA": row.get("iva"),
            "Tipo retención": display_rate(row.get("tipo_retencion", "")),
            "Retención": row.get("retencion"),
            "Suplidos": "",
            "Importe IRPF": row.get("retencion"),
            "Importe total sin suplidos": total,
            "Importe total": total,
            "Epígrafe": "",
        })
    return pd.DataFrame(records, columns=INGRESOS_COLUMNS)


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="276749")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="DDD6CE")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    headers = {cell.value: cell.column for cell in ws[1]}
    money_headers = {
        "Modelo 303 Base imponible", "Base imponible", "IVA", "Retención modelo 111", "Total",
        "Importe deducible IRPF", "Base IRPF", "Base imponible EUR", "BI Modelo 303",
        "Retención", "Suplidos", "Importe IRPF", "Importe total sin suplidos", "Importe total",
    }
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="E9E4DF"))
            header = ws.cell(row=1, column=cell.column).value
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00 €' if header in money_headers else '#,##0.00'
    for column in range(1, ws.max_column + 1):
        values = [str(ws.cell(row=r, column=column).value or "") for r in range(1, min(ws.max_row, 100) + 1)]
        width = min(max(len(value) for value in values) + 2, 42)
        ws.column_dimensions[get_column_letter(column)].width = max(10, width)
    ws.row_dimensions[1].height = 30


def build_excel(df: pd.DataFrame, details: list[dict[str, Any]], mode: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    if mode == "Recibidas (Gastos)":
        sheet.title = "Gastos"
        export_df = to_gastos(df)
    else:
        sheet.title = "Ingresos"
        export_df = to_ingresos(df)

    sheet.append(list(export_df.columns))
    for record in export_df.itertuples(index=False, name=None):
        sheet.append(list(record))
    _style_sheet(sheet)

    control = workbook.create_sheet("Control")
    control_headers = [
        "Archivo", "Concepto detectado", "Validación", "Confianza proveedor",
        "Confianza importes", "Fuente", "Observaciones",
    ]
    control.append(control_headers)
    for _, row in df.iterrows():
        control.append([
            row.get("archivo", ""),
            row.get("concepto", ""),
            row.get("validacion", ""),
            row.get("confianza_proveedor", ""),
            row.get("confianza_importes", ""),
            row.get("fuente_lectura", ""),
            row.get("error", ""),
        ])
    _style_sheet(control)

    if details:
        detail_sheet = workbook.create_sheet("Detalle IVA")
        detail_headers = ["Archivo", "Tipo IVA", "Base", "Cuota", "Texto detectado"]
        detail_sheet.append(detail_headers)
        for detail in details:
            detail_sheet.append([
                detail.get("archivo", ""), detail.get("tipo_iva", ""), detail.get("base"),
                detail.get("cuota"), detail.get("texto", ""),
            ])
        _style_sheet(detail_sheet)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
